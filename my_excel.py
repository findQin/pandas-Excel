import pandas as pd
from openpyxl import load_workbook
 
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
 
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
 
    writer = pd.ExcelWriter(filename, engine='openpyxl')
 
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
 
 
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
 
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
 
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
 
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
 
    if startrow is None:
        startrow = 0
 
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
 
    # save the workbook
    writer.save()



# 几个例子用于测试写入 excel 用的
df1 = pd.DataFrame({'One':[1,2,3], "Age": [22, 35, 58], "Sex": ["male", "male", "female"]})
df2 = pd.DataFrame({'Two':[4,5,6], "Age": [22, 35, 58], "Sex": ["male", "male", "female"]})
df3 = pd.DataFrame({'Three':[7,8,9], "Age": [22, 35, 58], "Sex": ["male", "male", "female"]})
df4 = pd.DataFrame({'Four':[10,11,12], "Age": [22, 35, 58], "Sex": ["male", "male", "female"]})



# 写入 Aho.xlsx 这个 excel 的 sheet1 表的第0行（startrow=0）,第0列（startcol=0）
append_df_to_excel('Aho.xlsx', df1, sheet_name='Sheet1', startcol=0,startrow=0,index=False)
 
# 写入sheet1 的第0行（startrow=0）,第3列（startcol=3）
append_df_to_excel('Aho.xlsx', df2, sheet_name='Sheet1', startcol=3,startrow=0,index=False)
 
# 写入sheet1 的第0行（startrow=0）,第7列（startcol=7）
append_df_to_excel('Aho.xlsx', df3, sheet_name='Sheet1', startcol=7,startrow=0,index=False)
 
# 写入sheet2 的第0行（startrow=0）,第0列（startcol=0）
append_df_to_excel('Aho.xlsx', df4, sheet_name='Sheet2', startcol=0,startrow=0,index=False)